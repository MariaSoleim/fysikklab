import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
import os
import math
from openpyxl import Workbook, load_workbook


g = 9.8214675  # +-0.0000004
m = 0.0027  # Fyll inn riktig masse
r = 0.02  # Fyll inn riktig radius
h = 0.01
inertia = 2/3*m*r**2


def plot_color(col, label):
    return mpatches.Patch(color=col, label=label)


Linear = plot_color('red', 'Lineær')
Sine = plot_color('blue', 'Sinus')
Steep = plot_color('green', 'Bratt')


colors = {'Linear': 'r-', 'Sine': 'b-', 'Steep': 'g-'}
euler_colors = {'Linear': 'r:', 'Sine': 'b:', 'Steep': 'g:'}


def euler(v, alpha):
    return v + ((0.03/5) * (g*math.sin(alpha)))


def trvalues(p, x):
    y = np.polyval(p, x)
    dp = np.polyder(p)
    dydx = np.polyval(dp, x)
    ddp = np.polyder(dp)
    d2ydx2 = np.polyval(ddp, x)
    alpha = np.arctan(-dydx)
    R = (1.0+dydx**2)**1.5/d2ydx2
    return [y, dydx, d2ydx2, alpha, R]


def get_alpha(poly, point):
    return trvalues(poly, point)[3]


def get_R(poly, point):
    return trvalues(poly, point)[4]


class Sheet:
    def __init__(self, sheet):
        self.sheet = sheet
        self.t = self.col('A')  # time
        self.x = self.col('B')  # x-position
        self.y = self.col('C')  # y-position

        # set up the sheets polynomial fit
        self.poly = self.iptrack()

        # compute the acceleration at each point
        self.acc = [self.compute_acceleration(x) for x in self.x]
        self.friction = [self.compute_friction(x) for x in self.x]
        self.normalforce = [self.compute_normal_power(x) for x in self.x]

        # for use by eulers method, initialize default values
        self.euler_velocity = [0]
        self.euler_dist = [self.x[0]]
        self.euler_acc = [self.compute_acceleration()]

    def col(self, col_name):
        return [float(c.value) for c in self.sheet[col_name]
                if c.value is not None]

    def iptrack(self):
        return np.polyfit(self.x, self.y, 15)

    def compute_normal_power(self, point=None):
        if point is None:
            # choose the first point as default initial value
            point = self.x[0]
        alpha = get_alpha(self.poly, point)
        return m * g * math.cos(alpha)

    def compute_friction(self, point=None):
        if point is None:
            # choose the first point as default initial value
            point = self.x[0]
        alpha = get_alpha(self.poly, point)
        return m * g * math.sin(alpha) + m * self.compute_acceleration(point)

    def compute_acceleration(self, point=None):
        if point is None:
            # choose the first point as default initial value
            point = self.x[0]
        alpha = get_alpha(self.poly, point)
        return g*math.sin(alpha)/(1+(inertia/m*r**2))

    """ NUMERICS """
    def euler_iteration(self):
        # fetch the first items in the experimental data sets
        x = self.euler_dist[-1]
        v = self.euler_velocity[-1]
        alpha = get_alpha(self.poly, x)

        # only estimate the next point if the x value
        # does not exceed the real limit
        x_next = x + h*v
        if x_next > max(self.x):
            x_next = x
        v_next = v + h*g*math.sin(alpha)/(1 + (inertia/m*r**2))
        a_next = self.compute_acceleration(x_next)

        self.euler_dist.append(x_next)
        self.euler_velocity.append(v_next)
        self.euler_acc.append(a_next)

    def compute_euler(self):
        for i in range(len(self.t)-1):
            self.euler_iteration()


wb = Workbook()
data_src = os.path.join(os.getcwd(), "lab2_data")
linear_book, sine_book, steep_book = [], [], []
data_dict = {"Linear": [], "Sine": [], "Steep": []}


def init():
    for curve_type in os.listdir(data_src):
        _path = os.path.join(data_src, curve_type)
        if "xlsx" not in _path:
            continue
        book = load_workbook(_path)
        sheets_found = 0
        for sheet in book.sheetnames:
            sheets_found += 1
            data_dict[curve_type[:-5]].append(Sheet(book[sheet]))
        print("found " + str(sheets_found) + " sheets!")


init()

for k, v in data_dict.items():
    print(k)
    # iterate over every xls document
    fig_num = 0
    euler_vals = {}  # a list of every point from each list (10)
    real_vals = {}  # decide what metrics to use here, dist/velo/acc
    # for each sheet!
    time_spent = []
    for s in v:
        s.compute_euler()
        plots = {
            'f_dist': zip(s.x, s.friction),
            'f_t': s.friction,
            'y_x': zip(s.x, s.y),
            'y_t': s.y,
            'acc_dist': zip(s.x, s.acc),
            'acc_t': s.acc,
            'N_dist': zip(s.x, s.normalforce),
            'N_t': s.normalforce,
            'EULER_acc_dist': zip(s.euler_dist, s.euler_acc),
            'EULER_acc_t': s.euler_acc,
            'EULER_dist': s.euler_dist
        }
        REAL_PLOT = 'acc_t'
        EULER_PLOT = 'EULER_acc_t'

        # TODO: set x,y data here
        current_time = 0
        for point in plots[REAL_PLOT]:
            try:
                real_vals[current_time].append(point)
            except KeyError:
                real_vals[current_time] = []
            current_time += 1

        current_time = 0
        for point in plots[EULER_PLOT]:
            try:
                euler_vals[current_time].append(point)
            except KeyError:
                euler_vals[current_time] = []
            current_time += 1

    """ PRUNE INVALID SETS (less than valid data) """
    invalid_euler_keys = []
    for w, e in euler_vals.items():
        if len(e) < 9:
            invalid_euler_keys.append(w)

    invalid_real_keys = []
    for w, e in real_vals.items():
        if len(e) < 9:
            invalid_real_keys.append(w)

    for key in invalid_euler_keys:
        euler_vals.pop(key, None)

    for key in invalid_real_keys:
        real_vals.pop(key, None)

    x_axis = []  # this is common between all plots
    euler_y_axis = []
    euler_y_axis_err = []
    y_axis = []
    y_axis_err = []
    for x_data, y_data in real_vals.items():
        if type(y_data[0]) == tuple:
            x_data = np.mean([val[0] for val in y_data])
            y_data = [val[1] for val in y_data]
        # do something with the data here!
        x_axis.append(x_data)
        variance = np.var(y_data)
        std = np.std(y_data)
        avg = np.mean(y_data)
        y_axis.append(avg)
        y_axis_err.append(std)

    for x_data, y_data in euler_vals.items():
        if type(y_data[0]) == tuple:
            x_data = np.mean([val[0] for val in y_data])
            y_data = [val[1] for val in y_data]
        variance = np.var(y_data)
        std = np.std(y_data)
        avg = np.mean(y_data)
        euler_y_axis.append(avg)
        euler_y_axis_err.append(std)
        print("var: " + str(variance) + ", std: " + str(std) + ", avg: " + str(avg))
    #  y_axis = y_axis[:len(x_axis)]
    y_axis_err = y_axis_err[:len(x_axis)]
    print('x,y: ' + str(len(x_axis)) + ', ' + str(len(y_axis)) )
    print('x,y euler: ' + str(len(x_axis)) + ', ' + str(len(euler_y_axis)) )
    #  plt.plot(x_axis, y_axis)
    plt.plot(x_axis, euler_y_axis, euler_colors[k])
    plt.errorbar(x_axis, y_axis, yerr=y_axis_err, fmt=colors[k])

# do something with s here
plt.legend(handles=[Linear, Sine, Steep])
plt.ylabel('akselerasjon (m/s^2)')
plt.xlabel('tid/ramme (100 fps)')
plt.show()
